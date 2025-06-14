#!/usr/bin/env python3
"""
Excel Quote Generator MVP
Drop in Excel file → Get formatted quote Excel out
"""

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import os
import json
import sys
from io import BytesIO
from datetime import datetime
from openai import OpenAI

def extract_customer_excel(file_path, images_output_dir):
    """Extract data and images from customer Excel file"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Build headers
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            headers.append(cell.value)
        else:
            headers.append(f"Unnamed_Column_{idx+1}")

    os.makedirs(images_output_dir, exist_ok=True)

    # Get worksheet dimensions
    max_data_row = ws.max_row
    
    print("=== EXTRACTING IMAGES ===")
    
    # First pass: calculate all image positions and scores for each row
    image_candidates = {}  # {image_index: [(row, score), ...]}
    
    for i, img in enumerate(ws._images):
        anchor = img.anchor
        from_row = anchor._from.row + 1  # Convert to 1-based
        from_col = anchor._from.col + 1
        
        # Handle different anchor types
        if hasattr(anchor, 'to') and anchor.to:
            to_row = anchor.to.row + 1
            to_col = anchor.to.col + 1
        else:
            # Use image size to estimate end position
            to_row = from_row + 2  # Assume 2 rows high by default
            to_col = from_col + 1
        
        print(f"Image {i}: from=({from_row},{from_col}) to=({to_row},{to_col})")
        
        # Calculate scores for all possible data rows
        candidates = []
        for data_row in range(2, max_data_row + 1):
            score = calculate_image_row_score(from_row, to_row, data_row)
            if score > 0:  # Only consider positive scores
                candidates.append((data_row, score))
        
        # Sort by score (highest first)
        candidates.sort(key=lambda x: x[1], reverse=True)
        image_candidates[i] = candidates

    # Second pass: use Hungarian-like assignment to avoid conflicts
    images_by_row = {}
    assigned_images = set()
    
    # Start with images that have fewer good options (more constrained)
    image_constraint_order = sorted(
        image_candidates.keys(), 
        key=lambda i: len([c for c in image_candidates[i] if c[1] > 5])
    )
    
    for img_idx in image_constraint_order:
        if img_idx in assigned_images:
            continue
            
        candidates = image_candidates[img_idx]
        
        # Find best unassigned row
        for row, score in candidates:
            if row not in images_by_row:
                images_by_row[row] = ws._images[img_idx]
                assigned_images.add(img_idx)
                print(f"Assigned image {img_idx} to row {row} (score: {score:.1f})")
                break
        else:
            # No unassigned row found, handle conflict
            if candidates:
                best_row, best_score = candidates[0]
                if best_row in images_by_row:
                    # Check if we should replace the existing assignment
                    current_img_idx = list(ws._images).index(images_by_row[best_row])
                    current_candidates = image_candidates.get(current_img_idx, [])
                    current_score = next((s for r, s in current_candidates if r == best_row), 0)
                    
                    if best_score > current_score * 1.2:  # 20% better to replace
                        # Try to relocate the current image
                        relocated = False
                        for alt_row, alt_score in current_candidates:
                            if alt_row not in images_by_row and alt_score > current_score * 0.7:  # Accept 30% score drop
                                images_by_row[alt_row] = images_by_row[best_row]
                                relocated = True
                                break
                        
                        if relocated or len(current_candidates) <= 1:  # Replace if relocated or current has no alternatives
                            images_by_row[best_row] = ws._images[img_idx]
                            assigned_images.add(img_idx)

    # Third pass: handle any remaining unassigned images
    unassigned_images = set(range(len(ws._images))) - assigned_images
    unassigned_rows = set(range(2, max_data_row + 1)) - set(images_by_row.keys())
    
    if unassigned_images and unassigned_rows:
        print(f"Assigning remaining {len(unassigned_images)} images to {len(unassigned_rows)} rows")
        
        # Simple proximity-based assignment for remainders
        for img_idx in unassigned_images:
            if not unassigned_rows:
                break
                
            img = ws._images[img_idx]
            anchor = img.anchor
            from_row = anchor._from.row + 1
            
            # Find closest unassigned row
            closest_row = min(unassigned_rows, key=lambda r: abs(r - from_row))
            images_by_row[closest_row] = img
            unassigned_rows.remove(closest_row)

    # Build the structured data
    structured_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                column_name = headers[col_idx]
                row_data[column_name] = cell.value if cell.value is not None else "null"

        image_filename = None
        if row_idx in images_by_row:
            openpyxl_img: OpenpyxlImage = images_by_row[row_idx]
            image_filename = f"image_row_{row_idx}.png"
            image_path = os.path.join(images_output_dir, image_filename)
            img_bytes = openpyxl_img._data()
            img_pil = PILImage.open(BytesIO(img_bytes))
            img_pil.save(image_path)

        row_data['image_file'] = image_filename if image_filename is not None else "null"
        structured_data.append(row_data)

    print(f"✅ Extracted {len(structured_data)} rows with {len(images_by_row)} images")
    return structured_data

def calculate_image_row_score(img_from_row, img_to_row, data_row):
    """Calculate how well an image matches a data row"""
    
    # Ensure img_to_row is at least img_from_row
    img_to_row = max(img_to_row, img_from_row)
    
    # Calculate overlap between image span and data row
    overlap_start = max(img_from_row, data_row)
    overlap_end = min(img_to_row, data_row + 1)
    overlap = max(0, overlap_end - overlap_start)
    
    if overlap <= 0:
        # No overlap, but calculate proximity penalty
        if data_row < img_from_row:
            distance = img_from_row - data_row
        else:
            distance = data_row - img_to_row
        
        # Proximity score (decreases with distance)
        if distance <= 2:
            return max(0, 2 - distance)  # Score 2 for adjacent, 1 for 1 row away
        else:
            return 0
    
    # Base score from overlap amount
    overlap_score = overlap * 10
    
    # Bonus for perfect alignment (image starts exactly at data row)
    alignment_bonus = 3 if img_from_row == data_row else 0
    
    # Bonus for image center being close to row center
    img_center = (img_from_row + img_to_row) / 2
    row_center = data_row + 0.5
    center_distance = abs(img_center - row_center)
    center_bonus = max(0, 2 - center_distance)  # Up to 2 points for perfect centering
    
    # Penalty for very large images (they might span multiple rows incorrectly)
    img_height = img_to_row - img_from_row
    size_penalty = min(2, max(0, img_height - 3))  # Penalty for images > 3 rows tall
    
    total_score = overlap_score + alignment_bonus + center_bonus - size_penalty
    
    return max(0, total_score)

def process_with_qwen(extracted_data):
    """Process extracted data using Qwen API"""
    
    # Check for API key
    if not os.environ.get("DASHSCOPE_API_KEY"):
        print("❌ DASHSCOPE_API_KEY environment variable not set")
        return None
    
    # Initialize OpenAI client for Qwen
    client = OpenAI(
        api_key=os.environ.get("DASHSCOPE_API_KEY"),
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )
    
    # Create prompt
    prompt = f"""You must translate and restructure the extracted JSON data into our internal CNC machining company Excel log format:

    Internal format required:
    - Serial_Number: 
    - Part_Name: 
    - Quantity: 
    - Material: 
    - Machining_Process: 
    - Surface_Finish: 
    - Notes: (or N/A if none)
    - image_file: (exact file name or null if not present - DO NOT modify filenames)

    Instructions to follow:
    - Translate Chinese keys exactly according to my mapping.
    - Keep original data values the same.
    - DON'T modify or lose the "image_file" reference. Pass it along unchanged.
    - Output structured valid JSON ONLY. NO OTHER TEXT. Don't output empty roles. 

    Here is the input JSON data:

    {json.dumps(extracted_data, ensure_ascii=False)}"""

    try:
        print("🤖 Processing with Qwen API...")
        response = client.chat.completions.create(
            model="qwen-turbo",
            messages=[{"role": "user", "content": prompt}],
        )
        
        # Get the response content
        qwen_response = response.choices[0].message.content.strip()
        
        # Clean and parse the response
        if qwen_response.startswith('```json'):
            qwen_response = qwen_response[7:-3]
        elif qwen_response.startswith('```'):
            qwen_response = qwen_response[3:-3]
        
        # Parse JSON response
        processed_data = json.loads(qwen_response)
        print(f"✅ Processed {len(processed_data)} items with Qwen")
        return processed_data
        
    except json.JSONDecodeError as e:
        print(f"❌ Error parsing JSON: {e}")
        print(f"Raw response: {qwen_response}")
        return None
    except Exception as e:
        print(f"❌ API call failed: {e}")
        return None

def generate_quote_excel(processed_data, output_filename):
    """Generate beautifully formatted modern quote Excel file (no yellow fills!)"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    # Modern color palette
    BRAND_BLUE = "2E86AB"      # Professional blue
    ACCENT_BLUE = "A23B72"     # Accent color
    LIGHT_GRAY = "F8F9FA"      # Very light background
    MEDIUM_GRAY = "6C757D"     # Text gray
    DARK_GRAY = "343A40"       # Dark text
    SUCCESS_GREEN = "28A745"   # For totals
    # FILL_YELLOW = "FFF3CD"   # Highlight color (REMOVED, not used)

    # Set default font for the entire sheet
    ws.sheet_properties.defaultRowHeight = 18
    
    # Modern header section with clean spacing
    ws.merge_cells('A1:H1')
    ws['A1'] = "手板报价单"
    ws['A1'].font = Font(name='Microsoft YaHei', size=28, bold=True, color=DARK_GRAY)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    # Party A (Customer) section - fillable (NO yellow fill)
    ws.merge_cells('A3:H3')
    ws['A3'] = "甲方信息"
    ws['A3'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=DARK_GRAY)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 25
    
    # Fillable customer fields (NO yellow)
    customer_fields = [
        "甲方公司: ___________________________",
        "联系人: ___________________________", 
        "电话: ___________________________",
        "邮箱: ___________________________"
    ]
    
    for i, field in enumerate(customer_fields, start=4):
        ws.merge_cells(f'A{i}:H{i}')
        ws[f'A{i}'] = field
        ws[f'A{i}'].font = Font(name='Microsoft YaHei', size=11, color=DARK_GRAY)
        ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')
        # NO fill: clean background only
        ws.row_dimensions[i].height = 22
    
    # Party B (Our company) section
    ws.merge_cells('A9:H9')
    ws['A9'] = "乙方信息"
    ws['A9'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=DARK_GRAY)
    ws['A9'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[9].height = 25
    
    # Company details in a clean layout
    company_info = [
        "乙方公司: 杭州越依模型科技有限公司",
        "联系人: 傅士勤",
        "电话: 137 7747 9066", 
        "地址: 杭州市富阳区东洲工业功能区1号路11号"
    ]
    
    for i, info in enumerate(company_info, start=10):
        ws.merge_cells(f'A{i}:H{i}')
        ws[f'A{i}'] = info
        ws[f'A{i}'].font = Font(name='Microsoft YaHei', size=11, color=DARK_GRAY)
        ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[i].height = 20

    # Add some breathing room
    ws.row_dimensions[14].height = 25
    
    # Modern table headers with clean design
    table_headers = ["序号", "零件图片", "零件名称", "表面处理", "材质", "数量", "单价(未税)", "总价(未税)"]
    header_widths = [6, 12, 20, 12, 15, 8, 15, 15]
    
    # Create header row with modern styling
    header_row = 15
    for col_num, (header, width) in enumerate(zip(table_headers, header_widths), 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # No borders for cleaner look
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[header_row].height = 35

    # Add data rows with alternating background and clean styling
    data_start_row = 16
    
    for idx, row_data in enumerate(processed_data):
        current_row = data_start_row + idx
        
        # Alternating row colors for better readability
        row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if idx % 2 == 0 else None
        
        # Set row height for images
        ws.row_dimensions[current_row].height = 65
        
        # Clean surface finish display
        surface_finish = row_data.get("Surface_Finish", "")
        if surface_finish is None or surface_finish == "null":
            surface_display = "—"
        else:
            surface_display = str(surface_finish).replace("120#", "").replace("+", " + ")
        
        # Extract and format quantity
        quantity = row_data.get("Quantity", 0)
        try:
            if isinstance(quantity, str) and quantity.strip():
                quantity = float(quantity)
            elif not isinstance(quantity, (int, float)):
                quantity = 0
        except (ValueError, TypeError):
            quantity = 0
        
        # Clean data values
        row_values = [
            f"{idx + 1:02d}",  # Zero-padded serial number
            "",  # Image placeholder
            row_data.get("Part_Name", "—"),
            surface_display,
            row_data.get("Material", "—"),
            int(quantity) if quantity == int(quantity) else quantity,
            0,  # Unit price placeholder
            None,  # Total will be formula
        ]
        
        # Write data with modern formatting
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_num)
            
            # Apply alternating row background
            if row_fill:
                cell.fill = row_fill
            
            # Special handling for total price column
            if col_num == 8:  # Total price column
                cell.value = f"=F{current_row}*G{current_row}"
                cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color=DARK_GRAY)
                cell.number_format = '"¥"#,##0.00'  # Currency formatting
            elif col_num == 7:  # Unit price column
                cell.value = value
                cell.font = Font(name='Microsoft YaHei', size=11, color=DARK_GRAY)
                cell.number_format = '"¥"#,##0.00'  # Currency formatting
                # NO yellow highlight
            elif col_num == 6:  # Quantity column
                cell.value = value
                cell.font = Font(name='Microsoft YaHei', size=11, color=DARK_GRAY)
                cell.number_format = '#,##0'  # Number formatting
            else:
                cell.value = value
                cell.font = Font(name='Microsoft YaHei', size=11, color=DARK_GRAY)
            
            # Modern alignment - no borders for cleaner look
            if col_num in [1, 2, 6, 7, 8]:  # Center align numbers and images
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Handle image insertion with better positioning
        if row_data.get("image_file") and row_data["image_file"] != "null":
            image_path = os.path.join("extracted_images", row_data["image_file"])
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    # Resize image to fit nicely in cell
                    max_size = 55
                    if img.width > max_size or img.height > max_size:
                        ratio = min(max_size/img.width, max_size/img.height)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    
                    # Center the image in the cell
                    img.anchor = f"B{current_row}"
                    ws.add_image(img)
                    
                except Exception as e:
                    print(f"⚠️  Error adding image for row {current_row}: {e}")

    # Modern totals section
    total_row = len(processed_data) + data_start_row + 1
    ws.row_dimensions[total_row].height = 35
    
    # Subtotal row
    ws.merge_cells(f'A{total_row}:G{total_row}')
    ws[f'A{total_row}'] = "小计"
    ws[f'A{total_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=DARK_GRAY)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
    
    first_data_row = data_start_row
    last_data_row = len(processed_data) + data_start_row - 1
    ws[f'H{total_row}'] = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws[f'H{total_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=SUCCESS_GREEN)
    ws[f'H{total_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'H{total_row}'].number_format = '"¥"#,##0.00'
    ws[f'H{total_row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    # Terms section with modern layout
    terms_start = total_row + 3
    
    # Terms header
    ws.merge_cells(f'A{terms_start}:H{terms_start}')
    ws[f'A{terms_start}'] = "条款说明"
    ws[f'A{terms_start}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=DARK_GRAY)
    ws[f'A{terms_start}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[terms_start].height = 25
    
    # Clean terms list with fillable delivery time (NO yellow highlight)
    terms = [
        "• 付款方式: 月结30天",
        "• 交货期: 确认后 (     ) 个工作日内完成",  # Fillable, but NO highlight
        "• 验收标准: 依据甲方2D、3D图纸及说明文档进行验收",
        "• 本报价单适用于杭州海康威视科技有限公司及其子公司、关联公司",
        "• 报价有效期: 30天",
        "• 所有价格均为人民币不含税价格"
    ]
    
    for i, term in enumerate(terms, start=terms_start + 1):
        ws.merge_cells(f'A{i}:H{i}')
        ws[f'A{i}'] = term
        ws[f'A{i}'].font = Font(name='Microsoft YaHei', size=10, color=MEDIUM_GRAY)
        ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[i].height = 20
        # NO yellow fill on any row

    # Modern signature section
    signature_row = terms_start + len(terms) + 3
    ws.merge_cells(f'F{signature_row}:H{signature_row}')
    ws[f'F{signature_row}'] = "乙方签名确认"
    ws[f'F{signature_row}'].font = Font(name='Microsoft YaHei', size=11, bold=True, color=DARK_GRAY)
    ws[f'F{signature_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[signature_row].height = 25
    
    # Add signature line
    ws.merge_cells(f'F{signature_row + 2}:H{signature_row + 2}')
    ws[f'F{signature_row + 2}'] = "________________________"
    ws[f'F{signature_row + 2}'].font = Font(name='Microsoft YaHei', size=10, color=MEDIUM_GRAY)
    ws[f'F{signature_row + 2}'].alignment = Alignment(horizontal='center', vertical='center')

    # Remove gridlines for cleaner look
    ws.sheet_view.showGridLines = False
    
    # Set page margins for better printing
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0
    
    # Save the workbook
    wb.save(output_filename)
    print(f"✅ Modern quote Excel generated: {output_filename}")
    return output_filename
    
     
def main():
    """Main pipeline function"""
    
    # Check command line arguments
    if len(sys.argv) < 2:
        print("Usage: python quote_generator.py <input_excel_file>")
        print("Example: python quote_generator.py input.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        sys.exit(1)
    
    print(f"🚀 Starting Excel Quote Generator")
    print(f"📁 Input file: {input_file}")
    
    try:
        # Step 1: Extract data and images from customer Excel
        print("\n=== STEP 1: EXTRACTING DATA ===")
        extracted_data = extract_customer_excel(input_file, "extracted_images")
        
        if not extracted_data:
            print("❌ No data extracted from Excel file")
            sys.exit(1)
        
        # Step 2: Process with Qwen API
        print("\n=== STEP 2: PROCESSING WITH QWEN ===")
        processed_data = process_with_qwen(extracted_data)
        
        if not processed_data:
            print("❌ Failed to process data with Qwen API")
            sys.exit(1)
        
        # Step 3: Generate quote Excel
        print("\n=== STEP 3: GENERATING QUOTE ===")
        output_filename = "手板报价单.xlsx"  # Simplified filename without date
        
        quote_file = generate_quote_excel(processed_data, output_filename)
        
        print(f"\n🎉 SUCCESS!")
        print(f"📊 Processed {len(processed_data)} items")
        print(f"📝 Quote generated: {quote_file}")
        
    except Exception as e:
        print(f"❌ Pipeline failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()