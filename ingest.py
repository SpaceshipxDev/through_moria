import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import os
from io import BytesIO

def extract_customer_excel(file_path, images_output_dir):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Build headers
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            headers.append(cell.value)
        else:
            headers.append(f"Unnamed_Column_{idx+1}")

    os.makedirs(images_output_dir, exist_ok=True)

    # Get worksheet dimensions
    max_data_row = ws.max_row
    
    print("=== IMPROVED IMAGE MAPPING ===")
    
    # First pass: calculate all image positions and scores for each row
    image_candidates = {}  # {image_index: [(row, score), ...]}
    
    for i, img in enumerate(ws._images):
        anchor = img.anchor
        from_row = anchor._from.row + 1  # Convert to 1-based
        from_col = anchor._from.col + 1
        
        # Handle different anchor types
        if hasattr(anchor, 'to') and anchor.to:
            to_row = anchor.to.row + 1
            to_col = anchor.to.col + 1
        else:
            # Use image size to estimate end position
            to_row = from_row + 2  # Assume 2 rows high by default
            to_col = from_col + 1
        
        print(f"Image {i}: from=({from_row},{from_col}) to=({to_row},{to_col})")
        
        # Calculate scores for all possible data rows
        candidates = []
        for data_row in range(2, max_data_row + 1):
            score = calculate_image_row_score(from_row, to_row, data_row)
            if score > 0:  # Only consider positive scores
                candidates.append((data_row, score))
        
        # Sort by score (highest first)
        candidates.sort(key=lambda x: x[1], reverse=True)
        image_candidates[i] = candidates
        
        print(f"  Top candidates: {candidates[:3]}")
    
    # Second pass: use Hungarian-like assignment to avoid conflicts
    images_by_row = {}
    assigned_images = set()
    
    # Start with images that have fewer good options (more constrained)
    image_constraint_order = sorted(
        image_candidates.keys(), 
        key=lambda i: len([c for c in image_candidates[i] if c[1] > 5])
    )
    
    for img_idx in image_constraint_order:
        if img_idx in assigned_images:
            continue
            
        candidates = image_candidates[img_idx]
        
        # Find best unassigned row
        for row, score in candidates:
            if row not in images_by_row:
                images_by_row[row] = ws._images[img_idx]
                assigned_images.add(img_idx)
                print(f"Assigned image {img_idx} to row {row} (score: {score:.1f})")
                break
        else:
            # No unassigned row found, handle conflict
            if candidates:
                best_row, best_score = candidates[0]
                if best_row in images_by_row:
                    # Check if we should replace the existing assignment
                    current_img_idx = list(ws._images).index(images_by_row[best_row])
                    current_candidates = image_candidates.get(current_img_idx, [])
                    current_score = next((s for r, s in current_candidates if r == best_row), 0)
                    
                    print(f"Conflict at row {best_row}: img {img_idx} (score {best_score:.1f}) vs img {current_img_idx} (score {current_score:.1f})")
                    
                    if best_score > current_score * 1.2:  # 20% better to replace
                        # Try to relocate the current image
                        relocated = False
                        for alt_row, alt_score in current_candidates:
                            if alt_row not in images_by_row and alt_score > current_score * 0.7:  # Accept 30% score drop
                                images_by_row[alt_row] = images_by_row[best_row]
                                print(f"  Relocated img {current_img_idx} to row {alt_row} (score: {alt_score:.1f})")
                                relocated = True
                                break
                        
                        if relocated or len(current_candidates) <= 1:  # Replace if relocated or current has no alternatives
                            images_by_row[best_row] = ws._images[img_idx]
                            assigned_images.add(img_idx)
                            print(f"  Replaced with img {img_idx}")
    
    # Third pass: handle any remaining unassigned images
    unassigned_images = set(range(len(ws._images))) - assigned_images
    unassigned_rows = set(range(2, max_data_row + 1)) - set(images_by_row.keys())
    
    if unassigned_images and unassigned_rows:
        print(f"\nAssigning remaining {len(unassigned_images)} images to {len(unassigned_rows)} rows")
        
        # Simple proximity-based assignment for remainders
        for img_idx in unassigned_images:
            if not unassigned_rows:
                break
                
            img = ws._images[img_idx]
            anchor = img.anchor
            from_row = anchor._from.row + 1
            
            # Find closest unassigned row
            closest_row = min(unassigned_rows, key=lambda r: abs(r - from_row))
            images_by_row[closest_row] = img
            unassigned_rows.remove(closest_row)
            print(f"Proximity assignment: img {img_idx} -> row {closest_row}")

    # Build the structured data
    structured_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                column_name = headers[col_idx]
                # Patch: Nulls become "null"
                row_data[column_name] = cell.value if cell.value is not None else "null"

        image_filename = None
        if row_idx in images_by_row:
            openpyxl_img: OpenpyxlImage = images_by_row[row_idx]
            image_filename = f"image_row_{row_idx}.png"
            image_path = os.path.join(images_output_dir, image_filename)
            img_bytes = openpyxl_img._data()
            img_pil = PILImage.open(BytesIO(img_bytes))
            img_pil.save(image_path)

        # Patch: Nulls become "null"
        row_data['image_file'] = image_filename if image_filename is not None else "null"
        structured_data.append(row_data)

    # Summary
    print(f"\n=== FINAL SUMMARY ===")
    print(f"Total images: {len(ws._images)}")
    print(f"Total data rows: {len(structured_data)}")
    print(f"Images successfully mapped: {len(images_by_row)}")
    print(f"Rows without images: {len(structured_data) - len(images_by_row)}")
    
    return structured_data

def calculate_image_row_score(img_from_row, img_to_row, data_row):
    """Calculate how well an image matches a data row"""
    
    # Ensure img_to_row is at least img_from_row
    img_to_row = max(img_to_row, img_from_row)
    
    # Calculate overlap between image span and data row
    overlap_start = max(img_from_row, data_row)
    overlap_end = min(img_to_row, data_row + 1)
    overlap = max(0, overlap_end - overlap_start)
    
    if overlap <= 0:
        # No overlap, but calculate proximity penalty
        if data_row < img_from_row:
            distance = img_from_row - data_row
        else:
            distance = data_row - img_to_row
        
        # Proximity score (decreases with distance)
        if distance <= 2:
            return max(0, 2 - distance)  # Score 2 for adjacent, 1 for 1 row away
        else:
            return 0
    
    # Base score from overlap amount
    overlap_score = overlap * 10
    
    # Bonus for perfect alignment (image starts exactly at data row)
    alignment_bonus = 3 if img_from_row == data_row else 0
    
    # Bonus for image center being close to row center
    img_center = (img_from_row + img_to_row) / 2
    row_center = data_row + 0.5
    center_distance = abs(img_center - row_center)
    center_bonus = max(0, 2 - center_distance)  # Up to 2 points for perfect centering
    
    # Penalty for very large images (they might span multiple rows incorrectly)
    img_height = img_to_row - img_from_row
    size_penalty = min(2, max(0, img_height - 3))  # Penalty for images > 3 rows tall
    
    total_score = overlap_score + alignment_bonus + center_bonus - size_penalty
    
    return max(0, total_score)

if __name__ == "__main__":
    file_path = "data/20250528-探野T4-3D打印手板加工清单.xlsx"
    images_dir = "extracted_images"

    data_list = extract_customer_excel(file_path, images_dir)

    import json
    print(json.dumps(data_list, indent=4, ensure_ascii=False))