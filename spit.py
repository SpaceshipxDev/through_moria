import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os

images_dir = "extracted_images"
output_excel = "output.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Components Log"

headers = ["Component ID", "Description", "Material", "Image"]
ws.append(headers)

data_for_testing = [
    {"id": "CMP001", "desc": "Component 1", "material": "Steel", "image_filename": "image_row_3.png"},
    {"id": "CMP002", "desc": "Component 2", "material": "Aluminum", "image_filename": "image_row_4.png"},
    {"id": "CMP003", "desc": "Component 3", "material": "Titanium", "image_filename": "image_row_5.png"},
]

for row_num, item in enumerate(data_for_testing, start=2):
    ws.cell(row=row_num, column=1, value=item["id"])
    ws.cell(row=row_num, column=2, value=item["desc"])
    ws.cell(row=row_num, column=3, value=item["material"])

    image_path = os.path.join(images_dir, item["image_filename"])
    if os.path.exists(image_path):
        img = OpenpyxlImage(image_path)

        # --- DYNAMIC RESIZING SECTION ---
        # Convert pixels to Excel's row height and column width units
        # (Excel: 1 row height unit ≈ 0.75 points ≈ 1.33 px; 1 col width ≈ px/7)
        px_height = img.height
        px_width = img.width

        # Convert for Excel: (rounded for Excel's weird units)
        row_height = int(px_height * 0.75)    # Excel's row height is about 0.75 points per pixel
        col_width = round(px_width / 7.0, 2)  # Excel's col width is about px/7

        ws.row_dimensions[row_num].height = row_height
        ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width or 0, col_width)

        img_anchor = f'D{row_num}'
        ws.add_image(img, img_anchor)
    else:
        print(f"Image file not found: {image_path}")

wb.save(output_excel)
print("✅ Universal tidy Excel generated.")