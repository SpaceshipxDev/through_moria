import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import os

def extract_customer_excel(file_path, images_output_dir):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Step 1: Dynamically get column headers (no hardcode)
    headers = {}
    for col_num, cell in enumerate(ws[1], start=1):  # Excel first row as headers
        headers[col_num] = cell.value if cell.value else f"Column_{col_num}"

    # Create directory clearly for images
    os.makedirs(images_output_dir, exist_ok=True)

    # Mapping image positions to rows
    images_by_row = {}
    for img in ws._images:
        anchor = img.anchor._from.row + 1  # openpyxl zero-indexes, so row +1
        images_by_row[anchor] = img

    # Now clearly collecting each row data into clearly structured json format
    structured_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # skip header row
        row_data = {}

        # Dynamically copy all cell data clearly from Excel row!
        for cell in row:
            column_name = headers[cell.column]
            row_data[column_name] = cell.value

        # Clearly checking if image exists for this row
        image_filename = None
        if row_idx in images_by_row:
            openpyxl_img: OpenpyxlImage = images_by_row[row_idx]

            # Save image clearly defined
            image_filename = f"image_row_{row_idx}.png"
            image_path = os.path.join(images_output_dir, image_filename)

            # Image saving handled clearly here
            openpyxl_img_ref: PILImage.Image = openpyxl_img._data()
            openpyxl_img_ref.save(image_path)
        
        # Clearly add the image_file data key
        row_data['image_file'] = image_filename

        # Append clearly row data for returning as structured dictionary clearly
        structured_data.append(row_data)

    return structured_data

# Let's test run clearly explained:
if __name__ == "__main__":
    file_path = "20250528-探野T4-3D打印手板加工清单.xlsx"  # customer input Excel
    images_dir = "extracted_images"

    data_list = extract_customer_excel(file_path, images_dir)

    # Clearly viewing extracted structured data
    import json
    print(json.dumps(data_list, indent=4))