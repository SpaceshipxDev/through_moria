from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

# Gemini structured data
gemini_output_data = [
  {
    "Serial_Number": 1,
    "Part_Name": "电池仓",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_3.png"
  },
  {
    "Serial_Number": 2,
    "Part_Name": "电池盖",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_4.png"
  },
  {
    "Serial_Number": 3,
    "Part_Name": "前壳",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_5.png"
  },
  {
    "Serial_Number": 4,
    "Part_Name": "按键1",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_6.png"
  },
  {
    "Serial_Number": 5,
    "Part_Name": "按键2",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_7.png"
  },
  {
    "Serial_Number": 6,
    "Part_Name": "按键3",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_8.png"
  },
  {
    "Serial_Number": 7,
    "Part_Name": "按键板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_9.png"
  },
  {
    "Serial_Number": 8,
    "Part_Name": "电源板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_10.png"
  },
  {
    "Serial_Number": 9,
    "Part_Name": "负极簧片",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_11.png"
  },
  {
    "Serial_Number": 10,
    "Part_Name": "负极接线片",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_12.png"
  },
  {
    "Serial_Number": 11,
    "Part_Name": "接口板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_13.png"
  },
  {
    "Serial_Number": 12,
    "Part_Name": "激光装饰片",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_14.png"
  },
  {
    "Serial_Number": 13,
    "Part_Name": "镜头",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_15.png"
  },
  {
    "Serial_Number": 14,
    "Part_Name": "镜头转接件1",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_16.png"
  },
  {
    "Serial_Number": 15,
    "Part_Name": "镜头转接件2",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_17.png"
  },
  {
    "Serial_Number": 16,
    "Part_Name": "镜头转接件3",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_18.png"
  },
  {
    "Serial_Number": 17,
    "Part_Name": "机芯支架",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_19.png"
  },
  {
    "Serial_Number": 18,
    "Part_Name": "散热板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_20.png"
  },
  {
    "Serial_Number": 19,
    "Part_Name": "手柄",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_21.png"
  },
  {
    "Serial_Number": 20,
    "Part_Name": "探测器板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_22.png"
  },
  {
    "Serial_Number": 21,
    "Part_Name": "显示屏",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_23.png"
  },
  {
    "Serial_Number": 22,
    "Part_Name": "显示屏壳体",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_24.png"
  },
  {
    "Serial_Number": 23,
    "Part_Name": "右饰片",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_25.png"
  },
  {
    "Serial_Number": 24,
    "Part_Name": "左饰片",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_26.png"
  },
  {
    "Serial_Number": 25,
    "Part_Name": "装饰片1",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_27.png"
  },
  {
    "Serial_Number": 26,
    "Part_Name": "装饰片2",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_28.png"
  },
  {
    "Serial_Number": 27,
    "Part_Name": "转接座",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_29.png"
  },
  {
    "Serial_Number": 28,
    "Part_Name": "转轴座",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_30.png"
  },
  {
    "Serial_Number": 29,
    "Part_Name": "主板",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_31.png"
  },
  {
    "Serial_Number": 30,
    "Part_Name": "主壳体",
    "Quantity": 1,
    "Material": "树脂",
    "Machining_Process": "3D打印",
    "Surface_Finish": "无",
    "Notes": "N/A",
    "image_file": "image_row_32.png"
  }
]

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "CNC_Log"

# Define headers
headers = ["Serial Number", "Part Name", "Quantity", "Material",
           "Machining Process", "Surface Finish", "Notes", "Image"]

# Write headers with formatting
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

# Set column widths for better spacing
column_widths = {
    'A': 12,  # Serial Number
    'B': 20,  # Part Name
    'C': 10,  # Quantity
    'D': 15,  # Material
    'E': 18,  # Machining Process
    'F': 15,  # Surface Finish
    'G': 25,  # Notes
    'H': 35   # Image (wider for images)
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add data rows with proper formatting
for idx, row_data in enumerate(gemini_output_data, start=2):
    # Set row height to accommodate images properly
    ws.row_dimensions[idx].height = 120
    
    # Data to write
    row_values = [
        row_data["Serial_Number"],
        row_data["Part_Name"],
        row_data["Quantity"],
        row_data["Material"],
        row_data["Machining_Process"],
        row_data["Surface_Finish"],
        "" if row_data["Notes"] in ["null", None] else row_data["Notes"],
        ""  # Image placeholder
    ]
    
    # Write data with formatting
    for col_num, value in enumerate(row_values, 1):
        cell = ws.cell(row=idx, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(size=10)
        
        # Add borders
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Alternate row coloring for better readability
        if idx % 2 == 0:
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Handle image insertion with proper sizing and positioning
    if row_data["image_file"] and row_data["image_file"] != "null":
        image_path = os.path.join("extracted_images", row_data["image_file"])
        if os.path.exists(image_path):
            try:
                img = XLImage(image_path)
                
                # Resize image to fit well in cell with margins
                # Maximum dimensions while leaving margins
                max_width = 200  # pixels
                max_height = 100  # pixels
                
                # Calculate scaling to maintain aspect ratio
                if img.width > max_width or img.height > max_height:
                    width_ratio = max_width / img.width
                    height_ratio = max_height / img.height
                    ratio = min(width_ratio, height_ratio)
                    
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                
                # Position image with some margin from cell edges
                # Calculate cell position with offset for margins
                col_letter = 'H'
                cell_address = f"{col_letter}{idx}"
                
                # Add some offset to center the image in the cell
                img.anchor = cell_address
                
                # Add image to worksheet
                ws.add_image(img)
                
            except Exception as e:
                print(f"Error adding image for row {idx}: {e}")
                # Add "No Image" text if image fails
                ws[f"H{idx}"].value = "No Image"
                ws[f"H{idx}"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Add "No Image" for missing images
        ws[f"H{idx}"].value = "No Image Available"
        ws[f"H{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"H{idx}"].font = Font(italic=True, color="999999")

# Freeze the header row for better navigation
ws.freeze_panes = "A2"

# Auto-filter for the header row
ws.auto_filter.ref = ws.dimensions

# Save the workbook
wb.save("professional_CNC_log.xlsx")
print("✅ Professional CNC Excel file created successfully with proper formatting!")